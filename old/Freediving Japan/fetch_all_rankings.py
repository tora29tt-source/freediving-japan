#!/usr/bin/env python3
"""
AIDA Japan Rankings - Complete Historical Fetcher
==================================================
Fetches ALL Japan ranking data for ALL years (1993-current) from
AIDA International website and saves to:
  - all_rankings_data.json  (raw data)
  - AIDA_Japan_全ランキングデータ_<today>.xlsx  (formatted Excel)

Usage:
    python3 -m pip install requests openpyxl --user
    python3 fetch_all_rankings.py

For weekly 2026 updates only:
    python3 fetch_all_rankings.py --update-2026

Takes 30-60 min for full historical fetch.
"""
import re, json, sys, time, datetime, argparse
import requests

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False
    print("⚠  openpyxl not found. JSON only. Run: python3 -m pip install openpyxl --user")

BASE      = "https://www.aidainternational.org"
RANK_POST = f"{BASE}/Ranking/index.php"
RANK_GET  = f"{BASE}/Ranking/index.php"

NATIONALITY_JAPAN = "108"
CURRENT_YEAR      = datetime.date.today().year
ALL_YEARS         = [str(y) for y in range(CURRENT_YEAR, 1992, -1)]  # 2026 → 1993

DISCIPLINES = [
    ("STA",  "8"),
    ("DYN",  "6"),
    ("DYNB", "11"),
    ("DNF",  "7"),
    ("CWT",  "3"),
    ("CWTB", "12"),
    ("CNF",  "4"),
    ("FIM",  "5"),
]
DISC_OVERALL = "all"
GENDERS      = [("Male", "0"), ("Female", "1")]

HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; FreedivingJapan/1.0)",
    "Content-Type": "application/x-www-form-urlencoded",
    "Referer": f"{BASE}/Ranking/",
}

ROW_RE  = re.compile(r"<tr[^>]*>(.*?)</tr>", re.S | re.I)
CELL_RE = re.compile(r"<t[dh][^>]*>(.*?)</t[dh]>", re.S | re.I)

def strip(html):
    return re.sub(r"\s+", " ", re.sub(r"<[^>]+>", "", html)).strip()

def extract_name(raw):
    return re.sub(r"\s*\([^)]+\)\s*$", "", raw).strip()

def set_filter(session, disc_val, gender_val, year_val):
    payload = {
        "discipline": disc_val,
        "nationality": NATIONALITY_JAPAN,
        "continent": "",
        "gender": gender_val,
        "year": year_val,
        "apply": "",
    }
    r = session.post(RANK_POST, data=payload, headers=HEADERS, timeout=30, allow_redirects=True)
    r.raise_for_status()
    return r.text

def fetch_page(session, page_num):
    url = RANK_GET if page_num == 1 else f"{RANK_GET}?page={page_num}"
    r = session.get(url, headers=HEADERS, timeout=30)
    r.raise_for_status()
    return r.text

def parse_records_rows(html, year, disc_name, gender_name):
    results = []
    for row in ROW_RE.findall(html):
        cells = [strip(c) for c in CELL_RE.findall(row)]
        if len(cells) < 7 or cells[0] == "#":
            continue
        try:
            rank = int(re.sub(r"\D", "", cells[0]))
        except ValueError:
            continue
        result = re.sub(r"\s*(NR|WR|CR|AR)\s*", "", cells[2]).strip()
        try:
            points = float(cells[4])
        except (ValueError, IndexError):
            points = None
        results.append({
            "year":       year,
            "discipline": disc_name,
            "gender":     gender_name,
            "rank":       rank,
            "name":       extract_name(cells[1]),
            "result":     result,
            "points":     points,
            "date":       cells[6] if len(cells) > 6 else "",
            "event":      cells[7] if len(cells) > 7 else "",
        })
    return results

def parse_overall_rows(html, year, gender_name):
    results = []
    disc_order = ["STA","DYN","DYNB","DNF","CWT","CWTB","CNF","FIM"]
    for row in ROW_RE.findall(html):
        cells = [strip(c) for c in CELL_RE.findall(row)]
        if len(cells) < 3 or cells[0] == "#":
            continue
        try:
            rank = int(re.sub(r"\D", "", cells[0]))
        except ValueError:
            continue
        try:
            total = float(cells[2])
        except ValueError:
            continue
        disc_pts = {}
        for i, dn in enumerate(disc_order):
            try:
                v = cells[3 + i]
                disc_pts[dn] = float(v) if v not in ("-", "") else None
            except (IndexError, ValueError):
                disc_pts[dn] = None
        results.append({
            "year":   year,
            "gender": gender_name,
            "rank":   rank,
            "name":   extract_name(cells[1]),
            "total":  str(total),
            **{f"disc_{k}": v for k, v in disc_pts.items()},
        })
    return results

def has_next_page(html, page):
    return f"page={page + 1}" in html

def scrape_combo(session, disc_val, disc_name, gender_val, gender_name, year):
    is_overall = (disc_val == DISC_OVERALL)
    html = set_filter(session, disc_val, gender_val, year)
    time.sleep(0.25)
    all_rows = []
    page = 1
    while True:
        rows = (parse_overall_rows if is_overall else parse_records_rows)(
            html, year, disc_name if not is_overall else "OVERALL", gender_name
        )
        all_rows.extend(rows)
        if not rows or not has_next_page(html, page):
            break
        page += 1
        html = fetch_page(session, page)
        time.sleep(0.25)
    return all_rows, page

# ── Excel output ────────────────────────────────────────────────────────────

BLUE   = "1565C0"
LBLUE  = "BBDEFB"
TEAL   = "00695C"
LTEAL  = "B2DFDB"
GREY   = "FAFAFA"
WHITE  = "FFFFFF"

def hdr_font(color="FFFFFF", bold=True):
    return Font(bold=bold, color=color, name="Arial", size=10)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)

def write_excel(all_records, all_overall, out_path):
    wb = openpyxl.Workbook()

    # ── Sheet 1: OVERALL ───────────────────────────────────────────────────
    ws_ov = wb.active
    ws_ov.title = "OVERALL"
    ov_hdrs = ["Year","Gender","Rank","Name","Total","STA","DYN","DYNB","DNF","CWT","CWTB","CNF","FIM"]
    for ci, h in enumerate(ov_hdrs, 1):
        c = ws_ov.cell(1, ci, h)
        c.font  = hdr_font()
        c.fill  = fill(BLUE)
        c.alignment = center()
        c.border = thin_border()
    ws_ov.row_dimensions[1].height = 20
    # Sort: year desc, gender, rank
    sorted_ov = sorted(all_overall, key=lambda r: (-int(r["year"]), r["gender"], r["rank"]))
    for ri, row in enumerate(sorted_ov, 2):
        bg = GREY if ri % 2 == 0 else WHITE
        vals = [
            row["year"], row["gender"], row["rank"], row["name"], row.get("total",""),
            row.get("disc_STA"), row.get("disc_DYN"), row.get("disc_DYNB"),
            row.get("disc_DNF"), row.get("disc_CWT"), row.get("disc_CWTB"),
            row.get("disc_CNF"), row.get("disc_FIM"),
        ]
        for ci, v in enumerate(vals, 1):
            c = ws_ov.cell(ri, ci, v)
            c.fill = fill(bg)
            c.border = thin_border()
            c.alignment = Alignment(horizontal="center" if ci != 4 else "left", vertical="center")
    # Column widths
    for ci, w in enumerate([6,8,6,22,8,7,7,7,7,7,7,7,7], 1):
        ws_ov.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 2: Records (per-discipline) ─────────────────────────────────
    ws_rc = wb.create_sheet("Records")
    rc_hdrs = ["Year","Discipline","Gender","Rank","Name","Result","Points","Date","Event"]
    for ci, h in enumerate(rc_hdrs, 1):
        c = ws_rc.cell(1, ci, h)
        c.font  = hdr_font()
        c.fill  = fill(TEAL)
        c.alignment = center()
        c.border = thin_border()
    ws_rc.row_dimensions[1].height = 20
    sorted_rc = sorted(all_records, key=lambda r: (-int(r["year"]), r["discipline"], r["gender"], r["rank"]))
    for ri, row in enumerate(sorted_rc, 2):
        bg = GREY if ri % 2 == 0 else WHITE
        vals = [row["year"], row["discipline"], row["gender"], row["rank"],
                row["name"], row["result"], row.get("points"), row["date"], row["event"]]
        for ci, v in enumerate(vals, 1):
            c = ws_rc.cell(ri, ci, v)
            c.fill = fill(bg)
            c.border = thin_border()
            c.alignment = Alignment(horizontal="center" if ci not in (5,9) else "left", vertical="center")
    for ci, w in enumerate([6,8,8,6,22,10,8,12,50], 1):
        ws_rc.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 3: Per-discipline summary sheets ─────────────────────────────
    for disc_name, _ in DISCIPLINES:
        ws = wb.create_sheet(disc_name)
        hdrs = ["Year","Gender","Rank","Name","Result","Points","Date","Event"]
        for ci, h in enumerate(hdrs, 1):
            c = ws.cell(1, ci, h)
            c.font = hdr_font()
            c.fill = fill(BLUE)
            c.alignment = center()
            c.border = thin_border()
        rows = [r for r in all_records if r["discipline"] == disc_name]
        rows = sorted(rows, key=lambda r: (-int(r["year"]), r["gender"], r["rank"]))
        for ri, row in enumerate(rows, 2):
            bg = GREY if ri % 2 == 0 else WHITE
            vals = [row["year"], row["gender"], row["rank"], row["name"],
                    row["result"], row.get("points"), row["date"], row["event"]]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(ri, ci, v)
                c.fill = fill(bg)
                c.border = thin_border()
                c.alignment = Alignment(horizontal="center" if ci not in (4,8) else "left", vertical="center")
        for ci, w in enumerate([6,8,6,22,10,8,12,50], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

    wb.save(out_path)
    print(f"  Saved Excel: {out_path}")

# ── Main ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--update-2026", action="store_true",
                        help="Only fetch/update current year data (fast, for weekly cron)")
    parser.add_argument("--years", nargs="+",
                        help="Override years to fetch, e.g. --years 2026 2022 2021")
    args = parser.parse_args()

    if args.update_2026:
        years = [str(CURRENT_YEAR)]
        print(f"🔄 Weekly update mode: fetching {CURRENT_YEAR} only")
    elif args.years:
        years = args.years
        print(f"📅 Fetching specified years: {years}")
    else:
        years = ALL_YEARS
        print(f"📚 Full historical fetch: {years[-1]}–{years[0]} ({len(years)} years)")

    session = requests.Session()
    session.get(f"{BASE}/Ranking/", headers=HEADERS, timeout=30)

    all_records, all_overall, errors = [], [], []
    total_combos = len(years) * len(GENDERS) * (len(DISCIPLINES) + 1)
    done = 0

    for year in years:
        print(f"\n=== {year} ===")
        has_any = False
        for gender_name, gender_val in GENDERS:
            for disc_name, disc_val in DISCIPLINES:
                try:
                    rows, pages = scrape_combo(session, disc_val, disc_name, gender_val, gender_name, year)
                    if rows:
                        all_records.extend(rows)
                        has_any = True
                    done += 1
                    prog = f"[{done}/{total_combos}]"
                    flag = "✓" if rows else "–"
                    print(f"  {prog} {flag} {year} {gender_name} {disc_name}: {len(rows)} athletes, {pages}p")
                except Exception as e:
                    print(f"  ✗ {year} {gender_name} {disc_name}: {e}")
                    errors.append(f"{year}/{gender_name}/{disc_name}: {e}")
                time.sleep(0.35)

            # OVERALL
            try:
                rows, pages = scrape_combo(session, DISC_OVERALL, "OVERALL", gender_val, gender_name, year)
                if rows:
                    all_overall.extend(rows)
                    has_any = True
                done += 1
                prog = f"[{done}/{total_combos}]"
                flag = "✓" if rows else "–"
                print(f"  {prog} {flag} {year} {gender_name} OVERALL: {len(rows)} athletes, {pages}p")
            except Exception as e:
                print(f"  ✗ {year} {gender_name} OVERALL: {e}")
                errors.append(f"{year}/{gender_name}/OVERALL: {e}")
            time.sleep(0.35)

        if not has_any:
            print(f"  (no Japan data for {year} — stopping early)")
            break

    # Save JSON
    today = datetime.date.today().isoformat()
    json_path = "all_rankings_data.json"
    out = {
        "updated": today,
        "years_fetched": years,
        "records": all_records,
        "overall": all_overall,
        "errors": errors,
    }
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=1)
    print(f"\n✅ JSON saved: {json_path}")

    # Save Excel
    if HAS_EXCEL:
        excel_path = f"AIDA_Japan_全ランキングデータ_{today}.xlsx"
        write_excel(all_records, all_overall, excel_path)
    else:
        print("⚠  Install openpyxl to get Excel output: python3 -m pip install openpyxl --user")

    print(f"\n📊 Summary")
    print(f"   Records (per discipline): {len(all_records)}")
    print(f"   Overall rankings:         {len(all_overall)}")
    if errors:
        print(f"   Errors: {len(errors)}")
        for e in errors: print(f"     • {e}")

if __name__ == "__main__":
    main()
